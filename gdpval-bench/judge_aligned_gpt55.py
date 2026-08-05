#!/usr/bin/env python3
"""Judge AA-aligned samples vs human. Model paths + judge endpoint from aligned.yaml.

Materials (closer to AA judge inputs; still single gpt-5.5 + cand-vs-human):
  - full task prompt (no 3000-char cut)
  - all reference files under tasks/<id>/refs/
  - submission/gold files with page images + extracted text
  - zip archives expanded and each member parsed

Usage:
  judge_aligned_gpt55.py [--config aligned.yaml]
  judge_aligned_gpt55.py --test [--config aligned.yaml]
"""
import os, sys, glob, json, base64, subprocess, tempfile, random, concurrent.futures, zipfile, shutil
sys.path.insert(0, os.path.expanduser("~/gdpval-bench"))
import aligned_config

_argv, _cfg_path = aligned_config.pop_config_args()
CFG = aligned_config.load(_cfg_path)
TR = os.path.join(CFG["_root"], "tasks")
MODEL = CFG["candidate"]["model"]
JUDGE_BASE = CFG["judge"]["base_url"]
JUDGE_KEY = CFG["judge"]["api_key"]
JUDGE_MODEL = CFG["judge"]["model"]
N_SAMPLES = int(CFG["judge"]["n_samples"])
MAXPAGES = int(CFG["judge"]["max_pages"])
WORKERS = int(CFG["judge"]["workers"])
# Caps so multimodal payloads stay practical while still much richer than before.
MAX_REF_FILES = int(CFG["judge"].get("max_ref_files", 16))
MAX_ZIP_MEMBERS = int(CFG["judge"].get("max_zip_members", 12))
MAX_IMGS_TOTAL = int(CFG["judge"].get("max_images_total", 48))
TEXT_CHARS = int(CFG["judge"].get("text_chars", 20000))
RESULTS_FILE = aligned_config.results_path(CFG, "judge")
DEVNULL = subprocess.DEVNULL
from openai import OpenAI
_cli = OpenAI(base_url=JUDGE_BASE, api_key=JUDGE_KEY, timeout=300, max_retries=3)

OFFICE_EXTS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xlsm", ".xls", ".odt", ".odp", ".ods"}
TEXT_EXTS = {".txt", ".md", ".csv", ".tsv", ".json", ".jsonl", ".py", ".js", ".ts", ".html", ".htm",
             ".xml", ".yaml", ".yml", ".log", ".rtf", ".tex", ".sql", ".r", ".ipynb"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}

JUDGE_INSTR = (
    "You are an expert occupational grader. You are given a work TASK, the TASK REFERENCE FILES "
    "provided to the worker, and two candidate deliverables A and B. Deliverables/references are "
    "shown as page images and/or extracted text (spreadsheets include cell/formula text; zips are "
    "expanded). Decide which deliverable better fulfills the task against the references, weighing "
    "accuracy, completeness, instruction-following and presentation. Be decisive. Respond ONLY "
    'with JSON: {"winner":"A" or "B" or "tie","reason":"<=25 words"}.'
)

def _ext(path):
    return os.path.splitext(path)[1].lower()

def b64(p):
    return "data:image/png;base64," + base64.b64encode(open(p, "rb").read()).decode()

def b64_image(p):
    ext = _ext(p).lstrip(".")
    mime = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "gif": "gif", "webp": "webp"}.get(ext, "png")
    return f"data:image/{mime};base64," + base64.b64encode(open(p, "rb").read()).decode()

def render_pages(src, wd, dpi=120, max_pages=None):
    max_pages = MAXPAGES if max_pages is None else max_pages
    os.makedirs(wd, exist_ok=True)
    ext = _ext(src)
    if ext == ".pdf":
        pdf = src
    elif ext in OFFICE_EXTS:
        prof = "file://" + tempfile.mkdtemp(prefix="lo_")
        try:
            subprocess.run(["soffice", "--headless", "-env:UserInstallation=" + prof, "--convert-to", "pdf",
                            "--outdir", wd, src], check=True, timeout=240, stdout=DEVNULL, stderr=DEVNULL)
        except Exception:
            return []
        pdf = os.path.join(wd, os.path.splitext(os.path.basename(src))[0] + ".pdf")
    else:
        return []
    if not os.path.exists(pdf):
        return []
    try:
        subprocess.run(["pdftoppm", "-png", "-r", str(dpi), "-l", str(max_pages), pdf, os.path.join(wd, "pg")],
                       check=True, timeout=240)
    except Exception:
        return []
    return sorted(glob.glob(wd + "/pg*.png"))

def xlsx_text(src, maxchars=None):
    maxchars = TEXT_CHARS if maxchars is None else maxchars
    if _ext(src) not in {".xlsx", ".xlsm", ".xls"}:
        return ""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(src, data_only=False)
    except Exception:
        return ""
    out = []
    for ws in wb.worksheets[:12]:
        out.append(f"# Sheet '{ws.title}' ({ws.max_row}x{ws.max_column})")
        for row in ws.iter_rows(max_row=min(ws.max_row or 0, 200), max_col=min(ws.max_column or 0, 40)):
            cells = [f"{c.coordinate}={c.value}" for c in row if c.value is not None]
            if cells:
                out.append(" ".join(cells))
    return "\n".join(out)[:maxchars]

def plain_text(src, maxchars=None):
    maxchars = TEXT_CHARS if maxchars is None else maxchars
    ext = _ext(src)
    if ext not in TEXT_EXTS:
        return ""
    try:
        raw = open(src, "rb").read()
    except Exception:
        return ""
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return raw.decode(enc)[:maxchars]
        except Exception:
            continue
    return raw.decode("utf-8", errors="replace")[:maxchars]

def docx_text(src, maxchars=None):
    maxchars = TEXT_CHARS if maxchars is None else maxchars
    if _ext(src) != ".docx":
        return ""
    try:
        import zipfile as zf
        from xml.etree import ElementTree as ET
        with zf.ZipFile(src) as z:
            xml = z.read("word/document.xml")
        root = ET.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        parts = [t.text for t in root.findall(".//w:t", ns) if t.text]
        return "\n".join(parts)[:maxchars]
    except Exception:
        return ""

def extract_text(src):
    return xlsx_text(src) or docx_text(src) or plain_text(src)

def expand_zip(src, out_dir):
    """Return list of extracted member file paths (non-dir), capped."""
    os.makedirs(out_dir, exist_ok=True)
    if _ext(src) != ".zip":
        return [src]
    try:
        with zipfile.ZipFile(src) as z:
            names = [n for n in z.namelist() if not n.endswith("/") and not n.startswith("__MACOSX/")]
            names = names[:MAX_ZIP_MEMBERS]
            paths = []
            for n in names:
                # flatten unsafe paths
                base = os.path.basename(n)
                if not base or base.startswith("."):
                    continue
                dst = os.path.join(out_dir, base)
                # disambiguate collisions
                i = 1
                stem, ext = os.path.splitext(base)
                while os.path.exists(dst):
                    dst = os.path.join(out_dir, f"{stem}_{i}{ext}")
                    i += 1
                with z.open(n) as srcf, open(dst, "wb") as outf:
                    shutil.copyfileobj(srcf, outf)
                paths.append(dst)
            return paths or [src]
    except Exception:
        return [src]

def materialize_file(src, cache_dir, label):
    """Return list of {label, path, imgs, text} blocks for one file (zip-expanded)."""
    blocks = []
    members = expand_zip(src, os.path.join(cache_dir, "unzip_" + label.replace("/", "_")))
    for i, path in enumerate(members):
        name = os.path.basename(path)
        tag = label if len(members) == 1 else f"{label}/{name}"
        wd = os.path.join(cache_dir, "render_" + tag.replace("/", "_").replace(" ", "_")[:80])
        imgs, text = [], ""
        ext = _ext(path)
        if ext in IMAGE_EXTS:
            imgs = [path]
            text = ""
        elif ext in OFFICE_EXTS or ext == ".pdf":
            imgs = render_pages(path, wd)
            text = extract_text(path)
        else:
            text = extract_text(path)
            if not text and ext not in {".mp4", ".mov", ".mp3", ".wav", ".m4a"}:
                # last resort: try office conversion anyway
                imgs = render_pages(path, wd)
        if ext in {".mp4", ".mov", ".mp3", ".wav", ".m4a"} and not imgs and not text:
            text = f"[binary media file present: {name}; content not inlined for this judge]"
        if imgs or text:
            blocks.append({"label": tag, "path": path, "imgs": imgs, "text": text})
    return blocks

def refs_blocks(idx):
    refs = sorted(glob.glob(f"{TR}/{idx}/refs/*"))
    refs = [p for p in refs if os.path.isfile(p)][:MAX_REF_FILES]
    cache = f"{TR}/{idx}/render_refs_full"
    out = []
    for p in refs:
        out.extend(materialize_file(p, cache, "ref:" + os.path.basename(p)))
    return out

def deliverable_blocks(path, cache_dir, label):
    return materialize_file(path, cache_dir, label)

def append_blocks(content, title, blocks, img_budget):
    """Append labeled materials; returns remaining image budget."""
    content.append({"type": "text", "text": title})
    if not blocks:
        content.append({"type": "text", "text": "(none)"})
        return img_budget
    for b in blocks:
        content.append({"type": "text", "text": f"--- {b['label']} ---"})
        for p in b["imgs"]:
            if img_budget <= 0:
                content.append({"type": "text", "text": "[additional pages omitted due to image budget]"})
                break
            try:
                content.append({"type": "image_url", "image_url": {"url": b64_image(p) if _ext(p) in IMAGE_EXTS else b64(p)}})
                img_budget -= 1
            except Exception:
                continue
        if b["text"]:
            content.append({"type": "text", "text": b["text"]})
    return img_budget

def judge(prompt, ref_blocks, a_blocks, b_blocks):
    content = [{"type": "text", "text": JUDGE_INSTR + "\n\nTASK:\n" + prompt}]
    budget = MAX_IMGS_TOTAL
    budget = append_blocks(content, "\n=== TASK REFERENCE FILES ===", ref_blocks, budget)
    # Split remaining budget roughly evenly across A/B.
    half = max(1, budget // 2)
    budget_a = append_blocks(content, "\n=== Deliverable A ===", a_blocks, half)
    # unused from A goes to B
    budget_b = budget - half + budget_a
    append_blocks(content, "\n=== Deliverable B ===", b_blocks, budget_b)
    content.append({"type": "text", "text": "Which deliverable is better? JSON only."})
    r = _cli.chat.completions.create(model=JUDGE_MODEL, messages=[{"role": "user", "content": content}], max_tokens=8000)
    raw = r.choices[0].message.content or ""
    try:
        j = json.loads(raw[raw.find("{"):raw.rfind("}") + 1])
        return j.get("winner"), j.get("reason", "")
    except Exception:
        return None, "parse-fail:" + raw[:120]

def gold_file(idx):
    g = glob.glob(f"{TR}/{idx}/gold/*")
    return g[0] if g else None

def sample_file(idx, s):
    d = glob.glob(f"{TR}/{idx}/cand_{MODEL}_aligned/s{s}/deliverable.*")
    return d[0] if d else None

def score_task(idx):
    gold = gold_file(idx)
    if not gold:
        return None
    g_blocks = deliverable_blocks(gold, f"{TR}/{idx}/render_gold_full", "human:" + os.path.basename(gold))
    if not g_blocks:
        return None
    prompt = open(f"{TR}/{idx}/prompt.txt").read()
    r_blocks = refs_blocks(idx)
    outcomes = []; reasons = []
    for s in range(N_SAMPLES):
        sf = sample_file(idx, s)
        if not sf:
            outcomes.append(0.0); reasons.append("no-deliverable")
            continue
        c_blocks = deliverable_blocks(sf, f"{TR}/{idx}/cand_{MODEL}_aligned/s{s}/render_full",
                                      "cand:" + os.path.basename(sf))
        if not c_blocks:
            outcomes.append(0.0); reasons.append("unrenderable"); continue
        pair = [("cand", c_blocks), ("human", g_blocks)]
        random.Random(idx * 10 + s).shuffle(pair)
        w, reason = judge(prompt, r_blocks, pair[0][1], pair[1][1])
        win = pair[0][0] if w == "A" else pair[1][0] if w == "B" else "tie"
        outcomes.append(1.0 if win == "cand" else 0.0 if win == "human" else 0.5)
        reasons.append(f"[{'cand' if win=='cand' else win} shownA={pair[0][0]}] {reason}")
    return (idx, sum(outcomes) / len(outcomes), outcomes, reasons)

def bootstrap_ci(vals, iters=2000):
    if not vals:
        return (0, 0)
    n = len(vals); rng = random.Random(13); m = []
    for _ in range(iters):
        m.append(sum(vals[rng.randrange(n)] for _ in range(n)) / n)
    m.sort(); return (m[int(0.025 * iters)], m[int(0.975 * iters)])

def main():
    tasks = aligned_config.tasks_list(CFG)
    print(f"aligned-judge config={CFG['_path']} model={MODEL} judge={JUDGE_MODEL}", flush=True)
    print(f"materials: full_prompt + refs(max {MAX_REF_FILES}) + zip_expand + "
          f"max_pages/doc={MAXPAGES} max_images={MAX_IMGS_TOTAL} text_chars={TEXT_CHARS}", flush=True)
    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=WORKERS) as ex:
        for i, r in enumerate(ex.map(score_task, tasks)):
            if r:
                results.append(r)
            if i % 25 == 0:
                print(f"  judged {i}/{len(tasks)}", flush=True)
    task_scores = [r[1] for r in results]
    wr = sum(task_scores) / len(task_scores) if task_scores else 0
    lo, hi = bootstrap_ci(task_scores)
    print(f"\n==== {MODEL} (materials-rich judge + {N_SAMPLES} samples) vs HUMAN ====")
    print(f"judge={JUDGE_MODEL} | cand-vs-human | non-production=loss | {len(results)} tasks")
    print(f"  win-rate vs human = {wr*100:.0f}%   (95% CI {lo*100:.0f}-{hi*100:.0f}%)")
    print(f"  results -> {RESULTS_FILE}", flush=True)
    json.dump([{"task": r[0], "score": r[1], "outcomes": r[2], "reasons": r[3]} for r in results],
              open(RESULTS_FILE, "w"), indent=1)

def test():
    idx = "0"
    gold = gold_file(idx)
    cand = glob.glob(f"{TR}/{idx}/cand_{MODEL}/deliverable.*")
    if not cand:
        cand = glob.glob(f"{TR}/{idx}/cand_{MODEL}_aligned/s*/deliverable.*")
    cand = cand[0] if cand else None
    print("gold:", os.path.basename(gold) if gold else None, "| cand:", os.path.basename(cand) if cand else None)
    if not gold or not cand:
        print("missing gold/cand for smoke test"); return
    prompt = open(f"{TR}/{idx}/prompt.txt").read()
    refs = refs_blocks(idx)
    g = deliverable_blocks(gold, "/tmp/rtest_g", "human")
    c = deliverable_blocks(cand, "/tmp/rtest_c", "cand")
    print(f"refs_blocks={len(refs)} gold_blocks={len(g)} cand_blocks={len(c)} prompt_chars={len(prompt)}")
    w, reason = judge(prompt, refs, c, g)
    print(f"{JUDGE_MODEL} verdict (A=cand,B=gold):", w, "|", reason)

if __name__ == "__main__":
    if "--test" in _argv:
        test()
    else:
        main()
